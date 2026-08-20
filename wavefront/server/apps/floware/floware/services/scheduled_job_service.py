import hashlib
import html
import io
import json
import os
import re
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from zoneinfo import ZoneInfo

from apscheduler.triggers.cron import CronTrigger
from common_module.log.logger import logger
from common_module.utils.serializer import serialize_values
from datasource import DatasourcePlugin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from db_repo_module.database.connection import DatabaseClient
from db_repo_module.models.datasource import Datasource
from db_repo_module.models.dynamic_query_yaml import DynamicQueryYaml
from db_repo_module.models.resource import ResourceScope
from db_repo_module.models.role import Role
from db_repo_module.models.scheduled_job import ScheduledJob
from db_repo_module.models.scheduled_job_execution import ScheduledJobExecution
from db_repo_module.models.user import User
from db_repo_module.models.user_role import UserRole
from db_repo_module.repositories.sql_alchemy_repository import SQLAlchemyRepository
from plugins_module.services.dynamic_query_service import DynamicQueryService
from plugins_module.services.datasource_services import (
    fetch_data_filters,
    get_datasource_config,
)
from sqlalchemy import select, text
from sqlalchemy.exc import IntegrityError
from user_management_module.constants.auth import SERVICE_AUTH_ROLE_ID
from user_management_module.services.email_service import EmailService
from user_management_module.services.user_service import UserService

STALE_LOCK_TIMEOUT_MINUTES = 30
SIGNED_URL_EXPIRY_SECONDS = 7 * 24 * 60 * 60
# Gmail and many providers cap attachments; keep below typical limits.
MAX_EMAIL_ATTACHMENT_BYTES = 10 * 1024 * 1024

COLUMN_FILL_COLORS = {
    'light_red': 'FFC7CE',
    'light_yellow': 'FFEB9C',
    'light_green': 'C6EFCE',
    'dark_green': '006100',
}
COLUMN_FILL_FONT_COLORS = {
    'dark_green': 'FFFFFF',
}

EMAIL_QUERY_PLACEHOLDER_PATTERN = re.compile(r'\{([a-zA-Z0-9_.\-]+)\}')


class ScheduledJobService:
    def __init__(
        self,
        db_client: DatabaseClient,
        scheduled_job_repository: SQLAlchemyRepository[ScheduledJob],
        scheduled_job_execution_repository: SQLAlchemyRepository[ScheduledJobExecution],
        datasource_repository: SQLAlchemyRepository[Datasource],
        dynamic_query_repository: SQLAlchemyRepository[DynamicQueryYaml],
        cloud_storage_manager,
        bucket_name: str,
        email_service: EmailService,
        user_repository: SQLAlchemyRepository[User],
        user_service: UserService,
        role_repository: SQLAlchemyRepository[Role],
        user_role_repository: SQLAlchemyRepository[UserRole],
    ):
        self.db_client = db_client
        self.scheduled_job_repository = scheduled_job_repository
        self.scheduled_job_execution_repository = scheduled_job_execution_repository
        self.datasource_repository = datasource_repository
        self.dynamic_query_repository = dynamic_query_repository
        self.cloud_storage_manager = cloud_storage_manager
        self.bucket_name = bucket_name
        self.email_service = email_service
        self.user_repository = user_repository
        self.user_service = user_service
        self.role_repository = role_repository
        self.user_role_repository = user_role_repository
        self.worker_id = os.getenv('HOSTNAME', 'floware-worker')
        self.dynamic_query_service = DynamicQueryService(
            cloud_storage_manager=self.cloud_storage_manager,
            dynamic_query_repo=self.dynamic_query_repository,
            bucket_name=self.bucket_name,
        )

    def _resolve_runtime_params(self, payload: dict, tz_name: str) -> dict | None:
        """Build query params for this run, including dynamic date presets."""
        base_params = payload.get('params')
        params: dict = dict(base_params) if isinstance(base_params, dict) else {}
        date_range = payload.get('date_range')
        if date_range not in {
            'last_day',
            't_2',
            'last_hour',
            'last_7_days',
            'last_30_days',
        }:
            return params or None

        start_key = payload.get('start_date_param', 'start_date')
        end_key = payload.get('end_date_param', 'end_date')

        # Time-window presets should be based on a stable clock. For `last_hour`,
        # use UTC explicitly (your DB timestamps are stored as UTC).
        if date_range == 'last_hour':
            end_dt_utc = datetime.now(timezone.utc).replace(microsecond=0)
            start_dt_utc = end_dt_utc - timedelta(hours=1)
            params[str(start_key)] = start_dt_utc.strftime('%Y-%m-%d %H:%M:%S')
            params[str(end_key)] = end_dt_utc.strftime('%Y-%m-%d %H:%M:%S')
            return params

        tz = ZoneInfo(tz_name)
        today = datetime.now(tz).date()
        if date_range == 'last_day':
            start_date = today - timedelta(days=1)
            end_date = today - timedelta(days=1)
        elif date_range == 't_2':
            # T-2: two calendar days before the run date in the job timezone.
            start_date = today - timedelta(days=2)
            end_date = today - timedelta(days=2)
        elif date_range == 'last_7_days':
            end_date = today - timedelta(days=1)
            start_date = end_date - timedelta(days=6)
        else:  # last_30_days
            end_date = today - timedelta(days=1)
            start_date = end_date - timedelta(days=29)

        params[str(start_key)] = start_date.isoformat()
        params[str(end_key)] = end_date.isoformat()
        return params

    def _compute_next_run_at(self, cron_expr: str, tz_name: str) -> datetime:
        tz = ZoneInfo(tz_name)
        now = datetime.now(tz)
        trigger = CronTrigger.from_crontab(cron_expr, timezone=tz)
        next_fire = trigger.get_next_fire_time(previous_fire_time=None, now=now)
        if next_fire is None:
            raise ValueError('Unable to compute next run time for cron expression')
        # Normalise to UTC-aware so the value written to the TIMESTAMPTZ column is
        # always unambiguous regardless of the PostgreSQL session timezone.
        return next_fire.astimezone(timezone.utc)

    async def create_job(
        self,
        job_type: str,
        cron_expr: str,
        timezone_name: str,
        payload: dict,
        max_retries: int,
    ) -> ScheduledJob:
        next_run_at = self._compute_next_run_at(cron_expr, timezone_name)
        return await self.scheduled_job_repository.create(
            job_type=job_type,
            cron_expr=cron_expr,
            timezone=timezone_name,
            payload=payload,
            next_run_at=next_run_at,
            status='active',
            max_retries=max_retries,
            retry_count=0,
        )

    async def list_jobs(
        self,
        limit: int = 100,
        offset: int = 0,
        job_type: str | None = None,
        status: str | None = None,
        payload_filters: dict[str, str] | None = None,
    ) -> list[ScheduledJob]:
        query = select(ScheduledJob)
        if job_type:
            query = query.where(ScheduledJob.job_type == job_type)
        if status:
            query = query.where(ScheduledJob.status == status)
        if payload_filters:
            for key, value in payload_filters.items():
                if key == 'query_id':
                    query = query.where(
                        text(
                            """
                            EXISTS (
                                SELECT 1
                                FROM jsonb_array_elements(
                                    scheduled_job.payload->'queries'
                                ) AS elem
                                WHERE elem->>'query_id' = :query_id
                            )
                            """
                        ).bindparams(query_id=value)
                    )
                    continue
                # JSONB subscript + astext for case-sensitive text comparison.
                query = query.where(ScheduledJob.payload[key].astext == value)
        query = (
            query.order_by(ScheduledJob.created_at.desc()).offset(offset).limit(limit)
        )
        async with self.db_client.session() as session:
            return (await session.scalars(query)).all()

    async def get_job(self, job_id: str) -> ScheduledJob | None:
        return await self.scheduled_job_repository.find_one(id=job_id)

    async def update_job(
        self,
        job_id: str,
        cron_expr: str | None = None,
        timezone_name: str | None = None,
        payload: dict | None = None,
        max_retries: int | None = None,
        status: str | None = None,
    ) -> ScheduledJob | None:
        updates = {}
        job = await self.scheduled_job_repository.find_one(id=job_id)
        if not job:
            return None

        effective_cron = cron_expr or job.cron_expr
        effective_tz = timezone_name or job.timezone
        if cron_expr or timezone_name:
            updates['next_run_at'] = self._compute_next_run_at(
                effective_cron, effective_tz
            )
            updates['cron_expr'] = effective_cron
            updates['timezone'] = effective_tz

        if payload is not None:
            updates['payload'] = payload
        if max_retries is not None:
            updates['max_retries'] = max_retries
        if status is not None:
            updates['status'] = status

        reactivating_failed = (
            job.status == 'failed'
            and updates
            and (status is None or status == 'active')
        )
        if reactivating_failed or status == 'active':
            updates['status'] = 'active'
            updates['retry_count'] = 0
            updates['last_error'] = None
        if reactivating_failed:
            updates['next_run_at'] = datetime.now(timezone.utc)

        if updates:
            return await self.scheduled_job_repository.find_one_and_update(
                filters={'id': job_id}, refresh=True, **updates
            )
        return job

    async def pause_job(self, job_id: str) -> ScheduledJob | None:
        return await self.scheduled_job_repository.find_one_and_update(
            filters={'id': job_id}, refresh=True, status='paused'
        )

    async def resume_job(self, job_id: str) -> ScheduledJob | None:
        job = await self.scheduled_job_repository.find_one(id=job_id)
        if not job:
            return None
        updates: dict = {
            'status': 'active',
            'retry_count': 0,
            'last_error': None,
        }
        if job.status == 'failed':
            # Run on the next poller tick rather than waiting until the next cron.
            updates['next_run_at'] = datetime.now(timezone.utc)
        else:
            updates['next_run_at'] = self._compute_next_run_at(
                job.cron_expr, job.timezone
            )
        return await self.scheduled_job_repository.find_one_and_update(
            filters={'id': job_id},
            refresh=True,
            **updates,
        )

    async def delete_job(self, job_id: str) -> bool:
        await self.scheduled_job_repository.delete_all(id=job_id)
        return True

    async def claim_due_jobs(self, batch_size: int = 10) -> list[dict]:
        claim_sql = text(
            """
            WITH due AS (
                SELECT id
                FROM scheduled_job
                WHERE status = 'active'
                  AND next_run_at <= now()
                ORDER BY next_run_at ASC
                FOR UPDATE SKIP LOCKED
                LIMIT :batch_size
            )
            UPDATE scheduled_job sj
            SET status = 'running',
                locked_by = :worker_id,
                locked_at = now(),
                updated_at = now()
            FROM due
            WHERE sj.id = due.id
            RETURNING sj.id, sj.job_type, sj.payload, sj.cron_expr, sj.timezone, sj.next_run_at, sj.retry_count, sj.max_retries
            """
        )
        async with self.db_client.session() as session:
            async with session.begin():
                res = await session.execute(
                    claim_sql,
                    {'batch_size': batch_size, 'worker_id': self.worker_id},
                )
                rows = [dict(row._mapping) for row in res.fetchall()]
            return rows

    async def _unlock_job(
        self,
        job_id: str,
        status: str,
        retry_count: int,
        last_error: str | None,
        next_run_at: datetime,
    ):
        """Unconditionally release the row-level lock on a job. Must never throw."""
        try:
            await self.scheduled_job_repository.find_one_and_update(
                filters={'id': job_id},
                refresh=False,
                status=status,
                retry_count=retry_count,
                last_error=last_error,
                last_run_at=datetime.now(timezone.utc),
                next_run_at=next_run_at,
                locked_by=None,
                locked_at=None,
            )
        except Exception as unlock_exc:
            logger.error(
                f'CRITICAL: failed to unlock job {job_id} after execution: {unlock_exc}'
            )

    async def _create_execution_lock(
        self, job_id: str, scheduled_for: datetime
    ) -> tuple[bool, str]:
        execution_key = f'{job_id}:{scheduled_for.isoformat()}'
        try:
            await self.scheduled_job_execution_repository.create(
                scheduled_job_id=job_id,
                execution_key=execution_key,
                scheduled_for=scheduled_for,
                status='running',
            )
            return True, execution_key
        except IntegrityError:
            # Unique constraint hit — another worker already claimed this fire time.
            logger.info(f'Skipping duplicate execution for job={job_id}')
            return False, execution_key
        except Exception as exc:
            # Real DB error — re-raise so the job is not silently skipped.
            logger.error(
                f'Unexpected error creating execution lock for job={job_id}: {exc}'
            )
            raise

    async def _complete_execution(
        self, execution_key: str, status: str, error: str | None = None
    ):
        try:
            await self.scheduled_job_execution_repository.find_one_and_update(
                filters={'execution_key': execution_key},
                status=status,
                error=error,
                refresh=False,
            )
        except Exception as exc:
            logger.error(
                f'Failed to update execution record key={execution_key}: {exc}'
            )

    async def recover_stale_locks(self):
        """Reset jobs stuck in 'running' and remove their orphaned in-progress execution locks.

        A crashed worker leaves the scheduled_job row in 'running' and the
        scheduled_job_execution row in 'running' (never completed).  If we only
        reset the job row the next claim hits the unique constraint on
        (scheduled_job_id, scheduled_for) and skips the fire.  Both cleanups
        must happen atomically.
        """
        recover_sql = text(
            """
            WITH stale AS (
                UPDATE scheduled_job
                SET status = 'active',
                    locked_by = NULL,
                    locked_at = NULL,
                    updated_at = now()
                WHERE status = 'running'
                  AND locked_at <= now() - make_interval(mins => :timeout_minutes)
                RETURNING id
            ),
            cleaned AS (
                DELETE FROM scheduled_job_execution
                WHERE scheduled_job_id IN (SELECT id FROM stale)
                  AND status = 'running'
                RETURNING scheduled_job_id
            )
            SELECT
                (SELECT count(*) FROM stale)   AS recovered_jobs,
                (SELECT count(*) FROM cleaned) AS cleaned_executions
            """
        )
        try:
            async with self.db_client.session() as session:
                async with session.begin():
                    res = await session.execute(
                        recover_sql,
                        {'timeout_minutes': STALE_LOCK_TIMEOUT_MINUTES},
                    )
                    row = res.fetchone()
                    recovered = int(row.recovered_jobs) if row else 0
                    cleaned = int(row.cleaned_executions) if row else 0
            if recovered:
                logger.warning(
                    f'Recovered {recovered} stale locked job(s); '
                    f'removed {cleaned} orphaned execution lock(s)'
                )
        except Exception as exc:
            logger.error(f'Failed to recover stale locks: {exc}')

    def recover_stale_locks_sync(self):
        import asyncio

        asyncio.run(self.recover_stale_locks())

    @staticmethod
    def _normalize_query_specs(payload: dict) -> list[dict]:
        queries = payload.get('queries')
        if not isinstance(queries, list) or not queries:
            raise ValueError('payload must include a non-empty queries array')

        specs: list[dict] = []
        seen_query_ids: set[str] = set()
        for index, item in enumerate(queries):
            if not isinstance(item, dict):
                raise ValueError(f'payload.queries[{index}] must be an object')
            query_id = item.get('query_id')
            if not query_id or not str(query_id).strip():
                raise ValueError(
                    f'payload.queries[{index}] must include a non-empty query_id'
                )
            normalized_query_id = str(query_id).strip()
            if normalized_query_id in seen_query_ids:
                raise ValueError(
                    f'payload.queries[{index}] duplicates query_id '
                    f'{normalized_query_id!r}'
                )
            seen_query_ids.add(normalized_query_id)
            specs.append(item)
        return specs

    @classmethod
    def _merge_query_spec(cls, job_payload: dict, query_spec: dict) -> dict:
        """Merge per-query overrides with job-level defaults."""
        job_params = job_payload.get('params')
        query_params = query_spec.get('params')
        merged_params: dict | None = None
        if isinstance(job_params, dict) or isinstance(query_params, dict):
            merged_params = {}
            if isinstance(job_params, dict):
                merged_params.update(job_params)
            if isinstance(query_params, dict):
                merged_params.update(query_params)

        return {
            'datasource_id': query_spec.get('datasource_id')
            or job_payload.get('datasource_id'),
            'query_id': str(query_spec['query_id']).strip(),
            'filter': query_spec.get('filter', job_payload.get('filter')),
            'offset': query_spec.get('offset', job_payload.get('offset', 0)),
            'limit': query_spec.get('limit', job_payload.get('limit', 100)),
            'column_styles': query_spec.get(
                'column_styles', job_payload.get('column_styles')
            ),
            'date_range': query_spec.get('date_range', job_payload.get('date_range')),
            'start_date_param': query_spec.get(
                'start_date_param', job_payload.get('start_date_param', 'start_date')
            ),
            'end_date_param': query_spec.get(
                'end_date_param', job_payload.get('end_date_param', 'end_date')
            ),
            'params': merged_params,
        }

    @staticmethod
    def _normalize_recipient_user_ids(payload: dict) -> list[str]:
        recipient_user_ids = payload.get('recipient_user_ids', [])
        if isinstance(recipient_user_ids, str):
            recipient_user_ids = [recipient_user_ids]
        return [
            str(user_id).strip()
            for user_id in recipient_user_ids
            if str(user_id).strip()
        ]

    async def _role_is_admin(self, role_id: str) -> bool:
        if role_id == SERVICE_AUTH_ROLE_ID:
            return True
        role = await self.role_repository.find_one(id=role_id)
        return bool(role and role.name == 'admin')

    async def _user_is_admin(self, user_id: str) -> bool:
        user_roles = await self.user_role_repository.find(user_id=user_id, limit=100)
        for user_role in user_roles:
            if await self._role_is_admin(str(user_role.role_id)):
                return True
        return False

    async def _rls_filter_for_user(self, user_id: str) -> str | None:
        if await self._user_is_admin(user_id):
            return None
        rls_filters = await self.user_service.get_user_resources(
            user_id=user_id, scope=ResourceScope.DATA
        )
        if len(rls_filters) == 0:
            raise ValueError(f'Data access not set for non-admin user {user_id}')
        additional_filters = fetch_data_filters(rls_filters)
        return f"{' $and '.join(additional_filters)}"

    async def _fetch_dynamic_query_rows(
        self,
        datasource_plugin: DatasourcePlugin,
        yaml_query: list,
        query_id: str,
        rls_filter_str: str | None,
        filter_expr: str | None,
        offset: int,
        limit: int,
        params: dict | None,
    ) -> list[dict]:
        result = await datasource_plugin.execute_dynamic_query(
            yaml_query,
            rls_filter_str,
            filter_expr,
            offset,
            limit,
            params,
        )
        result = serialize_values(result)
        if not result:
            raise ValueError(f'No results returned for dynamic query: {query_id}')
        first_key = next(iter(result))
        if result[first_key].get('status') != 'success':
            raise ValueError(
                f'Unexpected dynamic query result format for query_id {query_id}, '
                'no successful results'
            )
        rows = result[first_key].get('result') or []
        if not isinstance(rows, list):
            raise ValueError(
                f'Unexpected dynamic query result format for query_id {query_id}, invalid rows'
            )
        return rows

    @staticmethod
    def _sanitize_filename_part(value: str) -> str:
        return ''.join(
            ch if ch.isalnum() or ch in '-_' else '_' for ch in value.strip()
        )

    @classmethod
    def _build_report_filename(
        cls,
        *,
        query_id: str,
        user_id: str,
        run_timestamp: str,
        params: dict | None,
        payload: dict,
    ) -> str:
        unique_hash = hashlib.sha256(f'{user_id}:{run_timestamp}'.encode()).hexdigest()[
            :12
        ]
        start_key = str(payload.get('start_date_param', 'start_date'))
        end_key = str(payload.get('end_date_param', 'end_date'))
        start_date = params.get(start_key) if isinstance(params, dict) else None
        end_date = params.get(end_key) if isinstance(params, dict) else None

        if start_date and end_date:
            start_part = cls._sanitize_filename_part(str(start_date))
            end_part = cls._sanitize_filename_part(str(end_date))
            return f'{query_id}_{start_part}_{end_part}_{unique_hash}_report.xlsx'
        return f'{query_id}_{unique_hash}_report.xlsx'

    @staticmethod
    def _parse_numeric_cell_value(value) -> float | None:
        if value is None or value == '':
            return None
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        if isinstance(value, str):
            stripped = value.strip().replace(',', '')
            if not stripped:
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    @staticmethod
    def _normalize_column_key(name: str) -> str:
        return name.strip().casefold()

    @classmethod
    def _parse_column_styles_config(cls, raw_config) -> list[dict]:
        if not isinstance(raw_config, list):
            return []

        parsed_configs: list[dict] = []
        supported_ops = {'eq', 'neq', 'lt', 'lte', 'gt', 'gte', 'between'}
        for item in raw_config:
            if not isinstance(item, dict):
                continue
            column = item.get('column')
            rules = item.get('rules')
            if not isinstance(column, str) or not column.strip():
                continue
            if not isinstance(rules, list):
                continue

            parsed_rules: list[dict] = []
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                op = rule.get('op')
                fill = rule.get('fill')
                if (
                    op not in supported_ops
                    or not isinstance(fill, str)
                    or not fill.strip()
                ):
                    continue

                parsed_rule: dict = {'op': op, 'fill': fill.strip()}
                if op == 'between':
                    min_value = rule.get('min')
                    max_value = rule.get('max')
                    if not isinstance(min_value, (int, float)) or not isinstance(
                        max_value, (int, float)
                    ):
                        continue
                    parsed_rule['min'] = float(min_value)
                    parsed_rule['max'] = float(max_value)
                    parsed_rule['min_inclusive'] = bool(rule.get('min_inclusive', True))
                    parsed_rule['max_inclusive'] = bool(rule.get('max_inclusive', True))
                else:
                    value = rule.get('value')
                    if not isinstance(value, (int, float)):
                        continue
                    parsed_rule['value'] = float(value)
                parsed_rules.append(parsed_rule)

            if parsed_rules:
                parsed_configs.append({'column': column.strip(), 'rules': parsed_rules})
        return parsed_configs

    @staticmethod
    def _normalize_hex_color(value: str) -> str | None:
        if len(value) == 6 and all(ch in '0123456789ABCDEFabcdef' for ch in value):
            return value.upper()
        return None

    @classmethod
    def _resolve_fill_styles(
        cls, fill_name: str
    ) -> tuple[PatternFill | None, Font | None]:
        normalized = fill_name.strip()
        bg_hex: str | None = None
        font_hex: str | None = None

        if normalized.startswith('#'):
            bg_hex = cls._normalize_hex_color(normalized[1:])
        elif normalized.casefold() in COLUMN_FILL_COLORS:
            key = normalized.casefold()
            bg_hex = COLUMN_FILL_COLORS[key]
            font_hex = COLUMN_FILL_FONT_COLORS.get(key)
        else:
            bg_hex = cls._normalize_hex_color(normalized)

        if bg_hex is None:
            return None, None

        fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
        font = Font(color=font_hex) if font_hex else None
        return fill, font

    @classmethod
    def _rule_matches(cls, rule: dict, numeric_value: float) -> bool:
        op = rule['op']
        if op == 'eq':
            return numeric_value == rule['value']
        if op == 'neq':
            return numeric_value != rule['value']
        if op == 'lt':
            return numeric_value < rule['value']
        if op == 'lte':
            return numeric_value <= rule['value']
        if op == 'gt':
            return numeric_value > rule['value']
        if op == 'gte':
            return numeric_value >= rule['value']

        min_value = rule['min']
        max_value = rule['max']
        if rule.get('min_inclusive', True):
            if numeric_value < min_value:
                return False
        elif numeric_value <= min_value:
            return False
        if rule.get('max_inclusive', True):
            if numeric_value > max_value:
                return False
        elif numeric_value >= max_value:
            return False
        return True

    @classmethod
    def _build_column_style_map(
        cls, fieldnames: list[str], column_styles: list[dict]
    ) -> dict[int, list[dict]]:
        column_index = {
            cls._normalize_column_key(name): idx for idx, name in enumerate(fieldnames)
        }
        style_map: dict[int, list[dict]] = {}
        for config in column_styles:
            col_idx = column_index.get(cls._normalize_column_key(config['column']))
            if col_idx is None:
                logger.warning(
                    f'Column style config ignored; column not found: {config["column"]}'
                )
                continue
            style_map[col_idx] = config['rules']
        return style_map

    @classmethod
    def _apply_cell_style(cls, cell, raw_value, rules: list[dict]) -> None:
        numeric_value = cls._parse_numeric_cell_value(raw_value)
        if numeric_value is None:
            return
        for rule in rules:
            if not cls._rule_matches(rule, numeric_value):
                continue
            fill, font = cls._resolve_fill_styles(rule['fill'])
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            return

    @staticmethod
    def _fieldnames_from_rows(rows: list[dict]) -> list[str]:
        if not rows:
            return []
        fieldnames = list(rows[0].keys())
        for row in rows[1:]:
            for key in row:
                if key not in fieldnames:
                    fieldnames.append(key)
        return fieldnames

    @classmethod
    def _format_cell_display_value(cls, value) -> str:
        if isinstance(value, (dict, list)):
            return html.escape(json.dumps(value))
        if value is None:
            return ''
        return html.escape(str(value))

    @classmethod
    def _css_for_cell_value(cls, raw_value, rules: list[dict] | None) -> str:
        if not rules:
            return ''
        numeric_value = cls._parse_numeric_cell_value(raw_value)
        if numeric_value is None:
            return ''
        for rule in rules:
            if not cls._rule_matches(rule, numeric_value):
                continue
            fill, font = cls._resolve_fill_styles(rule['fill'])
            styles: list[str] = []
            if fill and fill.start_color:
                rgb = getattr(fill.start_color, 'rgb', None) or getattr(
                    fill.start_color, 'value', None
                )
                if rgb:
                    hex_color = str(rgb)[-6:]
                    styles.append(f'background-color:#{hex_color}')
            if font and font.color:
                rgb = getattr(font.color, 'rgb', None) or getattr(
                    font.color, 'value', None
                )
                if rgb:
                    hex_color = str(rgb)[-6:]
                    styles.append(f'color:#{hex_color}')
            if styles:
                return ';'.join(styles)
            return ''
        return ''

    @classmethod
    def _rows_to_html_table(
        cls,
        rows: list[dict],
        column_styles: list[dict] | None = None,
    ) -> str:
        if not rows:
            return '<p><em>No data</em></p>'

        fieldnames = cls._fieldnames_from_rows(rows)
        style_map = cls._build_column_style_map(fieldnames, column_styles or [])
        parts = [
            '<table border="1" cellpadding="6" cellspacing="0" '
            'style="border-collapse:collapse;margin:12px 0;max-width:100%;">',
            '<thead><tr>',
        ]
        for name in fieldnames:
            parts.append(
                f'<th style="font-weight:bold;background:#f3f3f3;text-align:left;">'
                f'{html.escape(name)}</th>'
            )
        parts.append('</tr></thead><tbody>')
        for row in rows:
            parts.append('<tr>')
            for col_idx, fieldname in enumerate(fieldnames):
                raw_value = row.get(fieldname)
                css = cls._css_for_cell_value(raw_value, style_map.get(col_idx))
                cell_html = cls._format_cell_display_value(raw_value)
                if css:
                    parts.append(f'<td style="{css}">{cell_html}</td>')
                else:
                    parts.append(f'<td>{cell_html}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table>')
        return ''.join(parts)

    @classmethod
    def _plain_text_to_html(cls, text: str) -> str:
        stripped = text.strip()
        if not stripped:
            return ''
        escaped = html.escape(stripped)
        paragraphs = escaped.split('\n\n')
        return ''.join(
            f'<p>{paragraph.replace(chr(10), "<br>")}</p>'
            for paragraph in paragraphs
            if paragraph
        )

    @classmethod
    def _looks_like_html(cls, content: str) -> bool:
        return bool(re.search(r'</?[a-zA-Z][^>]*>', content))

    @classmethod
    def _download_report_placeholder_html(cls, report: dict) -> str:
        report_name = html.escape(str(report['report_name']))
        report_size = int(report['report_size'])
        report_url = report.get('report_url')
        placeholder = (
            f'<p><b>{report_name}</b> ({report_size:,} bytes) exceeds the email '
            f'attachment limit.'
        )
        if report_url:
            placeholder += (
                f' <a href="{report_url}" target="_blank" '
                'rel="noopener noreferrer">Download Excel report</a>'
            )
        return f'{placeholder}</p>'

    @classmethod
    def _referenced_query_ids(cls, email_content: str | None) -> set[str]:
        if not email_content or not email_content.strip():
            return set()
        return {
            match.group(1)
            for match in EMAIL_QUERY_PLACEHOLDER_PATTERN.finditer(email_content.strip())
        }

    @classmethod
    def _build_query_tables_for_email(
        cls,
        prepared_reports: list[dict],
        referenced_query_ids: set[str],
    ) -> dict[str, str]:
        if not referenced_query_ids:
            return {}

        query_tables: dict[str, str] = {}
        for report in prepared_reports:
            query_id = report['query_id']
            if query_id not in referenced_query_ids:
                continue
            if report['use_attachment']:
                query_tables[query_id] = cls._rows_to_html_table(
                    report['rows'], column_styles=report.get('column_styles')
                )
            else:
                query_tables[query_id] = cls._download_report_placeholder_html(report)
        return query_tables

    @classmethod
    def _placeholder_table_html(
        cls, query_id: str, query_tables: dict[str, str]
    ) -> str:
        if query_id in query_tables:
            return query_tables[query_id]
        return (
            f'<p><em>No data for query '
            f'<code>{html.escape(query_id)}</code></em></p>'
        )

    @classmethod
    def _render_email_template(cls, template: str, query_tables: dict[str, str]) -> str:
        if not EMAIL_QUERY_PLACEHOLDER_PATTERN.search(template):
            if cls._looks_like_html(template):
                return template
            return cls._plain_text_to_html(template)

        is_html = cls._looks_like_html(template)
        parts: list[str] = []
        last_end = 0
        for match in EMAIL_QUERY_PLACEHOLDER_PATTERN.finditer(template):
            static = template[last_end : match.start()]
            if static:
                parts.append(static if is_html else cls._plain_text_to_html(static))
            parts.append(cls._placeholder_table_html(match.group(1), query_tables))
            last_end = match.end()
        trailing = template[last_end:]
        if trailing:
            parts.append(trailing if is_html else cls._plain_text_to_html(trailing))
        return ''.join(parts)

    @classmethod
    def _rows_to_xlsx_bytes(
        cls,
        rows: list[dict],
        column_styles: list[dict] | None = None,
    ) -> tuple[bytes, list[str]]:
        fieldnames = cls._fieldnames_from_rows(rows)

        def cell_value(value):
            if isinstance(value, (dict, list)):
                return json.dumps(value)
            return value if value is None or isinstance(value, str) else str(value)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Report'
        style_map = cls._build_column_style_map(fieldnames, column_styles or [])
        if fieldnames:
            header_font = Font(bold=True)
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                header_cell = worksheet.cell(row=1, column=col_idx, value=fieldname)
                header_cell.font = header_font
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    raw_value = row.get(fieldname)
                    cell = worksheet.cell(
                        row=row_idx, column=col_idx, value=cell_value(raw_value)
                    )
                    rules = style_map.get(col_idx - 1)
                    if rules:
                        cls._apply_cell_style(cell, raw_value, rules)

        buf = io.BytesIO()
        workbook.save(buf)
        return buf.getvalue(), fieldnames

    def _build_report_email_body(
        self,
        *,
        report_names: list[str],
        email_content: str | None,
        query_tables: dict[str, str],
        download_reports: list[dict],
    ) -> str:
        if email_content and email_content.strip():
            body = self._render_email_template(email_content.strip(), query_tables)
        elif len(report_names) == 1:
            body = f'<p>Scheduled report: <b>{report_names[0]}</b></p>'
        else:
            names_html = ', '.join(f'<b>{name}</b>' for name in report_names)
            body = f'<p>Scheduled reports: {names_html}</p>'

        if download_reports:
            body += (
                f'<p><b>Delivery:</b> The following report(s) exceed the '
                f'{MAX_EMAIL_ATTACHMENT_BYTES // (1024 * 1024)} MB email attachment limit. '
                'Use the download links below instead of attachments.</p>'
                '<p>Links are secure and expire in 7 days.</p>'
            )
            for report in download_reports:
                report_name = report['report_name']
                report_size = report['report_size']
                report_url = report.get('report_url')
                body += f'<p><b>{report_name}</b> ({report_size:,} bytes)'
                if report_url:
                    body += (
                        f' — <a href="{report_url}" target="_blank" '
                        'rel="noopener noreferrer">Download Excel report</a>'
                    )
                body += '</p>'
        return body

    async def _get_datasource_plugin(
        self, datasource_id: str
    ) -> tuple[DatasourcePlugin, str, str]:
        datasource_type, datasource_config = await get_datasource_config(
            datasource_id=datasource_id,
            datasource_repository=self.datasource_repository,
        )
        if not datasource_type or not datasource_config:
            raise ValueError(f'Datasource not found: {datasource_id}')
        return (
            DatasourcePlugin(datasource_type, datasource_config),
            datasource_type,
            datasource_id,
        )

    async def _execute_email_dynamic_query_job(self, payload: dict, job_timezone: str):
        query_specs = self._normalize_query_specs(payload)
        recipient_user_ids = self._normalize_recipient_user_ids(payload)
        email_content = payload.get('email_content')
        if email_content is not None and not isinstance(email_content, str):
            email_content = None

        if not recipient_user_ids:
            raise ValueError('payload must include recipient_user_ids')

        merged_specs = [self._merge_query_spec(payload, spec) for spec in query_specs]
        default_datasource_id = payload.get('datasource_id')
        if not default_datasource_id and not all(
            spec.get('datasource_id') for spec in merged_specs
        ):
            raise ValueError(
                'payload must include datasource_id (job-level or on each query)'
            )

        query_ids = [spec['query_id'] for spec in merged_specs]
        default_subject = payload.get('subject')
        if not default_subject:
            if len(query_ids) == 1:
                default_subject = f'Scheduled Dynamic Query Report: {query_ids[0]}'
            else:
                default_subject = 'Scheduled Dynamic Query Reports'
        subject = default_subject

        datasource_plugins: dict[str, DatasourcePlugin] = {}
        yaml_by_query_id: dict[str, tuple[list, str | None]] = {}

        for spec in merged_specs:
            query_id = spec['query_id']
            datasource_id = spec.get('datasource_id')
            if not datasource_id:
                raise ValueError(
                    f'datasource_id required for query_id={query_id} '
                    '(set at job level or on the query entry)'
                )
            if datasource_id not in datasource_plugins:
                plugin, _, _ = await self._get_datasource_plugin(datasource_id)
                datasource_plugins[datasource_id] = plugin
            if query_id not in yaml_by_query_id:
                (
                    yaml_query,
                    yaml_name,
                ) = await self.dynamic_query_service.get_dynamic_yaml_query(query_id)
                if not yaml_query:
                    raise ValueError(f'Dynamic query not found: {query_id}')
                yaml_by_query_id[query_id] = (yaml_query, yaml_name)

        failed_recipient_user_ids: list[str] = []
        delivered_count = 0
        run_timestamp = datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')

        for user_id in recipient_user_ids:
            user = await self.user_repository.find_one(id=user_id)
            if not user or user.deleted:
                logger.warning(
                    f'Scheduled report skipped: user not found or deleted ({user_id})'
                )
                failed_recipient_user_ids.append(user_id)
                continue

            try:
                rls_filter_str = await self._rls_filter_for_user(user_id)
            except Exception as exc:
                logger.error(f'Scheduled report failed for user_id={user_id}: {exc}')
                failed_recipient_user_ids.append(user_id)
                continue

            prepared_reports: list[dict] = []
            report_names: list[str] = []
            user_query_failed = False

            for spec in merged_specs:
                query_id = spec['query_id']
                datasource_id = spec['datasource_id']
                yaml_query, yaml_name = yaml_by_query_id[query_id]
                datasource_plugin = datasource_plugins[datasource_id]
                filter_expr = spec.get('filter')
                offset = spec.get('offset', 0)
                limit = spec.get('limit', 100)
                params = self._resolve_runtime_params(spec, job_timezone)
                column_styles = self._parse_column_styles_config(
                    spec.get('column_styles')
                )

                try:
                    rows = await self._fetch_dynamic_query_rows(
                        datasource_plugin,
                        yaml_query,
                        query_id,
                        rls_filter_str,
                        filter_expr,
                        offset,
                        limit,
                        params,
                    )
                except Exception as exc:
                    logger.error(
                        f'Scheduled report failed for user_id={user_id}, '
                        f'query_id={query_id}: {exc}'
                    )
                    user_query_failed = True
                    break

                if len(rows) == 0:
                    start_key = str(spec.get('start_date_param', 'start_date'))
                    end_key = str(spec.get('end_date_param', 'end_date'))
                    applied_start = (
                        params.get(start_key) if isinstance(params, dict) else None
                    )
                    applied_end = (
                        params.get(end_key) if isinstance(params, dict) else None
                    )
                    logger.info(
                        f'No records for user_id={user_id}, query_id={query_id}; '
                        f'range {applied_start}..{applied_end} (keys: {start_key}, {end_key}). '
                        'Skipping this report.'
                    )
                    continue

                report_bytes, _ = self._rows_to_xlsx_bytes(
                    rows, column_styles=column_styles
                )
                report_size = len(report_bytes)
                report_filename = self._build_report_filename(
                    query_id=query_id,
                    user_id=user_id,
                    run_timestamp=run_timestamp,
                    params=params,
                    payload=spec,
                )
                report_name = yaml_name or query_id
                report_names.append(report_name)
                use_attachment = report_size <= MAX_EMAIL_ATTACHMENT_BYTES
                report_url: str | None = None
                if not use_attachment:
                    report_key = f'scheduled_query_reports/{query_id}/{report_filename}'
                    self.cloud_storage_manager.save_small_file(
                        file_content=report_bytes,
                        bucket_name=self.bucket_name,
                        key=report_key,
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )
                    report_url = self.cloud_storage_manager.generate_presigned_url(
                        bucket_name=self.bucket_name,
                        key=report_key,
                        type='GET',
                        expiresIn=SIGNED_URL_EXPIRY_SECONDS,
                    )

                prepared_reports.append(
                    {
                        'query_id': query_id,
                        'report_name': report_name,
                        'report_size': report_size,
                        'report_url': report_url,
                        'use_attachment': use_attachment,
                        'filename': report_filename,
                        'content_bytes': report_bytes,
                        'rows': rows,
                        'column_styles': column_styles,
                    }
                )

            if user_query_failed:
                failed_recipient_user_ids.append(user_id)
                continue

            if not prepared_reports:
                logger.info(
                    f'No reports with data for user_id={user_id}; skipping email.'
                )
                continue

            download_reports = [
                {
                    'query_id': r['query_id'],
                    'report_name': r['report_name'],
                    'report_size': r['report_size'],
                    'report_url': r['report_url'],
                }
                for r in prepared_reports
                if not r['use_attachment']
            ]
            referenced_query_ids = self._referenced_query_ids(email_content)
            query_tables = self._build_query_tables_for_email(
                prepared_reports, referenced_query_ids
            )
            if referenced_query_ids:
                download_reports = [
                    report
                    for report in download_reports
                    if report['query_id'] not in referenced_query_ids
                ]
            body = self._build_report_email_body(
                report_names=report_names,
                email_content=email_content,
                query_tables=query_tables,
                download_reports=download_reports,
            )
            attachments = [
                {
                    'filename': r['filename'],
                    'content_bytes': r['content_bytes'],
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
                for r in prepared_reports
                if r['use_attachment']
            ]
            is_sent = self.email_service.send_email(
                subject,
                body,
                user.email,
                attachments=attachments or None,
            )
            if not is_sent:
                failed_recipient_user_ids.append(user_id)
            else:
                delivered_count += 1

        if failed_recipient_user_ids:
            query_label = ', '.join(query_ids)
            if delivered_count == 0 and len(failed_recipient_user_ids) == len(
                recipient_user_ids
            ):
                raise ValueError(
                    f'Failed scheduled report for all {len(recipient_user_ids)} '
                    f'recipient user(s) for queries={query_label}'
                )
            logger.error(
                f'Partial delivery failure for queries={query_label}: '
                f'{len(failed_recipient_user_ids)}/{len(recipient_user_ids)} '
                f'recipient user(s) failed: '
                f'{", ".join(failed_recipient_user_ids)}'
            )

    async def _run_job(self, job_row: dict):
        job_id = str(job_row['id'])
        scheduled_for = job_row['next_run_at']
        acquired, execution_key = await self._create_execution_lock(
            job_id, scheduled_for
        )
        if not acquired:
            # Another worker already executed this fire time. Release the row-level
            # lock that claim_due_jobs placed on the row so the job is not stuck
            # in 'running' until stale-lock recovery fires (30 min later).
            # We also advance next_run_at so the same execution_key is never hit
            # again, breaking the otherwise-infinite claim → duplicate → recover loop.
            try:
                next_run_at = self._compute_next_run_at(
                    job_row['cron_expr'], job_row['timezone']
                )
            except Exception:
                next_run_at = datetime.now(timezone.utc) + timedelta(minutes=10)
            await self._unlock_job(
                job_id=job_id,
                status='active',
                retry_count=int(job_row['retry_count']),
                last_error=None,
                next_run_at=next_run_at,
            )
            return

        job_exc: Exception | None = None
        try:
            if job_row['job_type'] != 'email_dynamic_query':
                raise ValueError(f"Unsupported job_type: {job_row['job_type']}")
            await self._execute_email_dynamic_query_job(
                job_row['payload'], job_row['timezone']
            )
        except Exception as exc:
            job_exc = exc
            logger.error(f'Failed scheduled job {job_id}: {exc}')
        finally:
            # Guaranteed unlock — always runs even if the DB call above throws.
            retry_count = int(job_row['retry_count'])
            if job_exc is None:
                final_status = 'active'
                final_retry_count = 0
                final_error = None
            else:
                retry_count += 1
                final_retry_count = retry_count
                final_status = (
                    'active' if retry_count <= int(job_row['max_retries']) else 'failed'
                )
                final_error = str(job_exc)

            # Compute next_run_at safely; fall back to a 10-minute delay if cron is broken.
            try:
                next_run_at = self._compute_next_run_at(
                    job_row['cron_expr'], job_row['timezone']
                )
            except Exception as cron_exc:
                logger.error(
                    f'Could not compute next_run_at for job {job_id}: {cron_exc}'
                )
                next_run_at = datetime.now(timezone.utc) + timedelta(minutes=10)

            await self._unlock_job(
                job_id=job_id,
                status=final_status,
                retry_count=final_retry_count,
                last_error=final_error,
                next_run_at=next_run_at,
            )
            execution_status = 'success' if job_exc is None else 'failed'
            await self._complete_execution(
                execution_key, status=execution_status, error=final_error
            )

    async def process_due_jobs(self, batch_size: int = 10):
        due_jobs = await self.claim_due_jobs(batch_size=batch_size)
        for job_row in due_jobs:
            await self._run_job(job_row)

    def process_due_jobs_sync(self, batch_size: int = 10):
        import asyncio

        asyncio.run(self.process_due_jobs(batch_size=batch_size))
